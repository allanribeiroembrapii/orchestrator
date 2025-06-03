import os
import pandas as pd
from dotenv import load_dotenv
from office365_api.upload_files_sebrae import upload_files
from openpyxl import load_workbook
from shutil import copyfile
from copy import copy

#carregar .env
load_dotenv()
ROOT = os.getenv('ROOT')

#sys.path
STEP1 = os.path.abspath(os.path.join(ROOT, 'step_1_data_raw'))
STEP2 = os.path.abspath(os.path.join(ROOT, 'step_2_stage_area'))
STEP3 = os.path.abspath(os.path.join(ROOT, 'step_3_data_processed'))

# Definindo as categorias de porte e fonte
def classificar_porte(porte):
    if porte in ["Grande", "Média"]:
        return "empresas_maior_porte"
    elif porte in ["Pequena", "Microempresa"]:
        return "empresas_menor_porte"
    else:
        return "empresas_menor_porte"
    
def classificar_fonte(fonte):
    if fonte in ["EMBRAPII", "BNDES", "Fundo de Investimento"]:
        return "valor_embrapii2"
    elif fonte in ["Unidade EMBRAPII"]:
        return "valor_unidade_embrapii2"
    elif fonte in ["SEBRAE"]:
        return "valor_sebrae2"
    elif fonte in ["Empresa"]:
        return "valor_empresa2_nao_sebrae"
    elif fonte in ["Micro e Pequena Empresa"]:
        return "valor_empresas_apoiadas"
    elif fonte in ["Média e Grande Empresa"]:
        return "valor_outras_empresas"
    else:
        return "outra_fonte"

# Gerando a NOVA planilha geral
def gerar_planilha_geral(gerar_novo = False, enviar_pasta_sebrae = False):

    # lendo as planilhas
    port = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'portfolio.xlsx')))
    proj_emp = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'projetos_empresas.xlsx')))
    emp = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'informacoes_empresas.xlsx')))
    emp_cont = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'empresas_contratantes.xlsx')))
    # oni_companies = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'oni_companies.xlsx')))
    company = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'srinfo_company_company.xlsx')))
    ppi = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'pedidos_pi.xlsx')))
    me = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'macroentregas.xlsx')))
    source = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'srinfo_sebrae_sourceamount.xlsx')))
    municipios = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'ibge_municipios.xlsx')))
    ue = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'info_unidades_embrapii.xlsx')))
    ue_local = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'srinfo_unit.xlsx')))
    ue_cnpj = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'Contas Bancárias - UE.xlsx')))

    if gerar_novo == False:
        # lendo as planilhas anteriores
        anterior_port = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'sebrae_geral.xlsx')), sheet_name = 'BD_portfolio')
        anterior_ue = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'sebrae_geral.xlsx')), sheet_name = 'BD_UEs')
        anterior_proj_emp = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'sebrae_geral.xlsx')), sheet_name = 'BD_projetos_empresas')
        anterior_emp = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'sebrae_geral.xlsx')), sheet_name = 'BD_empresas')
        anterior_me = pd.read_excel(os.path.abspath(os.path.join(STEP1, 'sebrae_geral.xlsx')), sheet_name = 'BD_macroentregas')

        # mudando a planilha anterior para dado_atual = 0
        anterior_port['dado_atual'] = 0
        anterior_ue['dado_atual'] = 0
        anterior_proj_emp['dado_atual'] = 0
        anterior_emp['dado_atual'] = 0
        anterior_me['dado_atual'] = 0

    ## COLUNAS CALCULADAS

    # valor_total
    port['valor_total'] = port['valor_embrapii'] + port['valor_empresa'] + port['valor_unidade_embrapii'] + port['valor_sebrae']

    # pedidos_pi
    contagem_ppi = ppi.groupby('codigo_projeto').size().reset_index(name = 'pedidos_pi')
    port_pedidos = pd.merge(port, contagem_ppi, on = 'codigo_projeto', how = 'left')

    # macroentregas
    contagem_mes = me.groupby('codigo_projeto').size().reset_index(name = 'num_macroentregas')
    port_mes = pd.merge(port_pedidos, contagem_mes, on = 'codigo_projeto', how = 'left')

    # numero de MPEs e MGEs
    # juntando projetos_empresas e informacoes_empresas
    combinado = pd.merge(proj_emp, emp, on = 'cnpj', how = 'left')
    # contando numero de empresas por porte em cada projeto
    contagem_empresas = combinado.groupby(['codigo_projeto', 'porte']).size().reset_index(name = 'num_empresas')
    contagem_empresas["categoria"] = contagem_empresas["porte"].apply(classificar_porte)
    contagem_empresas = contagem_empresas.groupby(['codigo_projeto', 'categoria'])['num_empresas'].sum().unstack(fill_value=0).reset_index()
    port_empresas = pd.merge(port_mes, contagem_empresas, on = 'codigo_projeto', how = 'left')

    # valores por fonte (MPEs, MGEs, SEBRAE, EMBRAPII, Unidade EMBRAPII)
    source["categoria"] = source["fonte"].apply(classificar_fonte)
    source_grouped = source.groupby(['codigo_negociacao', 'categoria'])['valor'].sum().unstack(fill_value=0).reset_index()
    port_fonte = pd.merge(port_empresas, source_grouped, on = 'codigo_negociacao', how = 'left')



    ####### PLANILHA GERAL #######
    planilha_geral = port_fonte[port_fonte['parceria_programa'].str.contains('SEBRAE')]

    # incluindo coluna de id sequencial, dado_atual e data de extração
    if gerar_novo == False:
        planilha_geral['n'] = range(max(anterior_port['n']) + 1, max(anterior_port['n']) + len(planilha_geral) + 1)
    else:
        planilha_geral['n'] = range(1, len(planilha_geral) + 1)

    planilha_geral['dado_atual'] = 1
    planilha_geral['data_extracao'] = pd.to_datetime('today').date()

    # escolhendo as coluna, renomeando e reordenando
    novos_nomes_e_ordem = {
        'n': 'n',
        'dado_atual': 'dado_atual',
        'data_extracao': 'data_extracao',
        'codigo_projeto': 'codigo_projeto',
        'codigo_negociacao': 'codigo_negociacao',
        'unidade_embrapii': 'unidade_embrapii',
        'data_contrato': 'data_contrato',
        'data_termino': 'data_termino',
        'valor_total': 'valor_total',
        'titulo_publico': 'titulo_publico',
        'descricao_publica': 'descricao_publica',
        'tipo_projeto': 'tipo_projeto',
        'tecnologia_habilitadora': 'tecnologia_habilitadora',
        'macroentregas': 'num_macroentregas',
        'status': 'status',
        'pct_aceites': 'percentual_projeto',
        'trl_inicial': 'trl_inicial',
        'trl_final': 'trl_final',
        'valor_unidade_embrapii': 'valor_unidade_embrapii',
        'valor_empresas_apoiadas': 'valor_empresas_apoiadas',
        'valor_outras_empresas': 'valor_outras_empresas',
        'valor_sebrae': 'valor_sebrae',
        'valor_embrapii': 'valor_embrapii',
        'modalidade_financiamento': 'modalidade_financiamento',
        'empresas_menor_porte': 'empresas_menor_porte',
        'empresas_maior_porte': 'empresas_maior_porte',
        'pedidos_pi': 'pedidos_pi',
    }

    planilha_geral_final = planilha_geral[novos_nomes_e_ordem.keys()]
    planilha_geral_final = planilha_geral_final.rename(columns=novos_nomes_e_ordem)

    # ajustando coluna de data
    planilha_geral_final['data_contrato'] = pd.to_datetime(planilha_geral_final['data_contrato'], format='%d/%m/%Y', errors='coerce').dt.date
    planilha_geral_final['data_termino'] = pd.to_datetime(planilha_geral_final['data_termino'], format='%d/%m/%Y', errors='coerce').dt.date

    # ajustando coluna de percentual
    planilha_geral_final['percentual_projeto'] = planilha_geral_final['percentual_projeto'] * 100

    # ajustando coluna de modalidade
    planilha_geral_final['modalidade_financiamento'] = planilha_geral_final['modalidade_financiamento'].apply(
        lambda x: 'DT' if '(DT)' in x else ('ET' if '(ET)' in x else 'AT')
    )

    # juntando a planilha anterior com a nova
    if gerar_novo == False:
        planilha_geral_final = pd.concat([anterior_port, planilha_geral_final], ignore_index=True)



    ####### PLANILHA AUXILIAR DE UEs #######
    # organizando informações de CNPJ das UEs
    ue_cnpj = ue_cnpj[['UNIDADE', 'CNPJ (BANCO)']].drop_duplicates()
    ue_cnpj = ue_cnpj.rename(columns={
        'UNIDADE': 'unidade_embrapii',
        'CNPJ (BANCO)': 'cnpj'
    })

    # Dicionário de substituição
    substituicoes = {
        'BIOINDÚSTRIA E BIOINSUMOS - UFT': 'Bioindústria e Bioinsumos - UFT',
        'CIN-UFPE': 'CIn-UFPE',
        'CNPEM BIOMASSA': 'CNPEM',
        'CPQD': 'CPqD',
        'FITec CAMPINAS': 'FITec Campinas',
        'GRAPHENE-UCS': 'Graphene-UCS',
        'IF-GOIANO': 'IF-Goiano',
        'IF-GOIÁS': 'IF-Goiás',
        'IF-MG ': 'IF-MG',
        'IF-SULDEMINAS': 'IF-Suldeminas',
        'IPT-BIO': 'IPT-Bio',
        'IPT-MATERIAIS': 'IPT-Materiais',
        'LESC-UFC': 'LESC - UFC',
        'POLIMEROS': 'Polimeros',
        'SENAI ISI BIOSSINTÉTICO': 'SENAI ISI BIOSSINTÉTICOS',
        'SIMOB-UFRGS -': 'SIMOB-UFRGS',
        'UFV - FIBRAS FLORESTAIS ': 'UFV - Fibras Florestais',
        'VBL IOT E INDÚSTRIA 4.0 - Von Braun': 'VBL IoT E INDÚSTRIA 4.0 - Von Braun'
    }

    # Aplicando a substituição
    ue_cnpj['unidade_embrapii'] = ue_cnpj['unidade_embrapii'].replace(substituicoes)

    # juntando as planilhas
    ue = pd.merge(ue, ue_cnpj, on = 'unidade_embrapii', how = 'left')
    port_ue = pd.merge(planilha_geral, ue, on = 'unidade_embrapii', how = 'left')
    port_ue = pd.merge(port_ue, ue_local, left_on = 'unidade_embrapii', right_on = 'name', how = 'left')
    

    # retirando ";" de 'competencias_tecnicas'
    port_ue['competencias_tecnicas'] = port_ue['competencias_tecnicas'].str.replace(';', '')

    # Selecionar as colunas desejadas
    colunas_desejadas_ues = [
        'unidade_embrapii',
        'ue_responsavel_institucional',
        'tipo_instituicao',
        'zip_code',
        'address',
        'municipio',
        'uf',
        'competencias_tecnicas',
        'cnpj'
        ]   

    # agrupando por unidade e selecionando as colunas necessárias
    # Agrupando a coluna 'cnpj' concatenando com '; ' e mantendo a primeira ocorrência para as demais
    port_ue2 = port_ue[colunas_desejadas_ues].groupby('unidade_embrapii').agg({
        'ue_responsavel_institucional': 'first',
        'tipo_instituicao': 'first',
        'zip_code': 'first',
        'address': 'first',
        'municipio': 'first',
        'uf': 'first',
        'competencias_tecnicas': 'first',
        'cnpj': lambda x: '; '.join(x.dropna().unique())  # concatena CNPJs únicos separados por "; "
    }).reset_index()


    if gerar_novo == False:
        port_ue2['n'] = range(max(anterior_ue['n']) + 1, max(anterior_ue['n']) + len(port_ue2) + 1)
    else:
        port_ue2['n'] = range(1, len(port_ue2) + 1)

    port_ue2['dado_atual'] = 1
    port_ue2['data_extracao'] = pd.to_datetime('today').date()

    # renomeando as colunas
    port_ue2 = port_ue2.rename(columns={
        'unidade_embrapii': 'unidade_embrapii',
        'ue_responsavel_institucional': 'responsavel_institucional',
        'tipo_instituicao': 'tipo_instituicao',
        'zip_code': 'cep',
        'address': 'endereco',
        'municipio': 'municipio',
        'uf': 'uf',
        'competencias_tecnicas': 'competencias_tecnicas',
        'cnpj': 'cnpj_representante_financeiro'
    })

    colunas_desejadas_ues = [
        'n',
        'dado_atual',
        'data_extracao',
        'unidade_embrapii',
        'responsavel_institucional',
        'tipo_instituicao',
        'cep',
        'endereco',
        'municipio',
        'uf',
        'competencias_tecnicas',
        'cnpj_representante_financeiro'
    ]

    port_ue2 = port_ue2[colunas_desejadas_ues]
    
    # CEPs em branco - ajuste
    cep_substituicoes = {
        'BIOTEC-CETENE': '50740-545',
        'IF-ES': '29056-255',
    }

    # Aplicando as substituições
    for unidade, novo_cep in cep_substituicoes.items():
        port_ue2.loc[port_ue2['unidade_embrapii'] == unidade, 'cep'] = novo_cep

    # Endereço em branco - ajuste
    endereco_substituicoes = {
        'BIOTEC-CETENE': 'Av. Prof. Luiz Freire, 01',
        'IF-ES': ' Av. Rio Branco, 50, Santa Lúcia',
    }

    # Aplicando as substituições
    for unidade, novo_endereco in endereco_substituicoes.items():
        port_ue2.loc[port_ue2['unidade_embrapii'] == unidade, 'endereco'] = novo_endereco

    # juntando a planilha anterior com a nova
    if gerar_novo == False:
        port_ue2 = pd.concat([anterior_ue, port_ue2], ignore_index=True)



    ####### PLANILHA AUXILIAR PROJETOS EMPRESAS #######
    projetos_empresas = proj_emp[proj_emp['codigo_projeto'].isin(planilha_geral['codigo_projeto'])]
    if gerar_novo == False:
        projetos_empresas['n'] = range(max(anterior_proj_emp['n']) + 1, max(anterior_proj_emp['n']) + len(projetos_empresas) + 1)
    else:
        projetos_empresas['n'] = range(1, len(projetos_empresas) + 1)
    projetos_empresas['dado_atual'] = 1
    projetos_empresas['data_extracao'] = pd.to_datetime('today').date()

    colunas_desejadas_proj_emp = ['n', 'dado_atual', 'data_extracao', 'codigo_projeto', 'cnpj']
    projetos_empresas = projetos_empresas[colunas_desejadas_proj_emp]

    # juntando a planilha anterior com a nova
    if gerar_novo == False:
        projetos_empresas = pd.concat([anterior_proj_emp, projetos_empresas], ignore_index=True)



    ####### PLANILHA AUXILIAR DE EMPRESAS #######
    # pegando informações mais recentes de oni_companies
    # oni_companies = oni_companies[oni_companies['ativo'] == 1]

    # juntando as planilhas
    emp = pd.merge(emp, company[['cnpj', 'neighborhood', 'zip_code']], on = 'cnpj', how = 'left')
    proj_emp2 = pd.merge(proj_emp, emp, on = 'cnpj', how = 'left')
    proj_emp2 = pd.merge(proj_emp2, emp_cont, on = 'cnpj', how = 'left')
    port_emp = pd.merge(planilha_geral, proj_emp2, on = 'codigo_projeto', how = 'left')

    # Selecionar as colunas desejadas
    colunas_desejadas_emp = [
            'cnpj',
            'empresa_x',
            'contatos',
            'emails_contatos',
            'telefones_contatos',
            'zip_code',
            'neighborhood',
            'municipio',
            'uf',
            'porte',
            'faixa_faturamento_x'
            ]      

    # agrupando por empresa e selecionando as colunas desejadas
    port_emp2 = port_emp[colunas_desejadas_emp].groupby(['cnpj'], as_index=False).first()
    
    if gerar_novo == False:
        port_emp2['n'] = range(max(anterior_emp['n']) + 1, max(anterior_emp['n']) + len(port_emp2) + 1)
    else:
        port_emp2['n'] = range(1, len(port_emp2) + 1)
    port_emp2['dado_atual'] = 1
    port_emp2['data_extracao'] = pd.to_datetime('today').date()

    # renomeando as colunas
    port_emp2 = port_emp2.rename(columns={
        'cnpj': 'cnpj',
        'empresa_x': 'razao_social',
        'contatos': 'contato_declarado',
        'emails_contatos': 'e-mail_de_contato',
        'telefones_contatos': 'telefone',
        'zip_code': 'cep',
        'neighborhood': 'bairro',
        'municipio': 'municipio',
        'uf': 'uf',
        'porte': 'porte',
        'faixa_faturamento_x': 'faixa_faturamento'
        })
    
    colunas_desejadas_emp = [
        'n',
        'dado_atual',
        'data_extracao',
        'cnpj',
        'razao_social',
        'contato_declarado',
        'e-mail_de_contato',
        'telefone',
        'cep',
        'bairro',
        'municipio',
        'uf',
        'porte',
        'faixa_faturamento'
    ]

    port_emp2 = port_emp2[colunas_desejadas_emp]

    # juntando a planilha anterior com a nova
    if gerar_novo == False:
        port_emp2 = pd.concat([anterior_emp, port_emp2], ignore_index=True)
    


    ####### PLANILHA AUXILIAR DE MACROENTREGAS #######
    # juntando as planilhas
    port_me = pd.merge(planilha_geral['codigo_projeto'], me, on = 'codigo_projeto', how = 'left')

    # Selecionar as colunas desejadas
    colunas_desejadas_me = [
        'codigo_projeto',
        'num_macroentrega',
        'titulo',
        'descricao_macroentrega',
        'valor_embrapii_me',
        'valor_empresa_me',
        'valor_unidade_embrapii_me',
        'data_inicio_planejado',
        'data_termino_planejado',
        'data_inicio_real',
        'data_termino_real',
        'percentual_executado',
        'data_aceitacao'
        ]   

    # agrupando por projeto e macroentrega e selecionando as colunas desejadas
    port_me = port_me[colunas_desejadas_me].groupby(['codigo_projeto', 'num_macroentrega'], as_index=False).first()

    # calculando o valor total
    port_me['valor_total'] = port_me['valor_embrapii_me'] + port_me['valor_empresa_me'] + port_me['valor_unidade_embrapii_me']
    if gerar_novo == False:
        port_me['n'] = range(max(anterior_me['n']) + 1, max(anterior_me['n']) + len(port_me) + 1)
    else:
        port_me['n'] = range(1, len(port_me) + 1)
    port_me['dado_atual'] = 1
    port_me['data_extracao'] = pd.to_datetime('today').date()

    # Selecionar as colunas desejadas
    colunas_desejadas_me2 = [
        'n',
        'dado_atual',
        'data_extracao',
        'codigo_projeto',
        'num_macroentrega',
        'titulo',
        'descricao_macroentrega',
        'valor_total',
        'data_inicio_planejado',
        'data_termino_planejado',
        'data_inicio_real',
        'data_termino_real',
        'percentual_executado',
        'data_aceitacao'
        ]  
    
    # selecionando somente as desejadas
    port_me2 = port_me[colunas_desejadas_me2]

    # ajustando colunas de data
    campos_data = ['data_inicio_planejado', 'data_termino_planejado', 'data_inicio_real', 'data_termino_real', 'data_aceitacao']
    for campo in campos_data:
        if campo in port_me2.columns:
            port_me2[campo] = pd.to_datetime(port_me2[campo], format='%d/%m/%Y', errors='coerce').dt.date

    # juntando a planilha anterior com a nova
    if gerar_novo == False:
        port_me2 = pd.concat([anterior_me, port_me2], ignore_index=True)

    
    # Salvando o arquivo
    caminho_arquivo = os.path.abspath(os.path.join(STEP3, 'sebrae_geral.xlsx'))

    if not gerar_novo:
        # Copia o arquivo base
        origem = os.path.abspath(os.path.join(STEP1, 'sebrae_geral.xlsx'))
        copyfile(origem, caminho_arquivo)

        # Copia a aba Dicionário preservando a formatação
        copiar_planilha_formatada(origem, caminho_arquivo, 'Dicionário')

    # Escrevendo cada DataFrame em uma aba diferente
    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        planilha_geral_final.to_excel(writer, sheet_name='BD_portfolio', index=False)
        port_ue2.to_excel(writer, sheet_name='BD_UEs', index=False)
        projetos_empresas.to_excel(writer, sheet_name='BD_projetos_empresas', index=False)
        port_emp2.to_excel(writer, sheet_name='BD_empresas', index=False)
        port_me2.to_excel(writer, sheet_name='BD_macroentregas', index=False)

    # Garante que ao menos uma aba está visível
    wb = load_workbook(caminho_arquivo)
    abas_visiveis = [sheet for sheet in wb.worksheets if sheet.sheet_state == 'visible']
    if not abas_visiveis:
        wb.worksheets[0].sheet_state = 'visible'
    wb.save(caminho_arquivo)
    wb.close()

    if enviar_pasta_sebrae:
        upload_files(
            pasta_arquivos=STEP3,
            destino='Acompanhamento Descentralizado//base_de_dados_sebrae_nacional',
            arquivo_especifico=caminho_arquivo
        )

    return planilha_geral, combinado, municipios, port_ue, proj_emp, port_emp, port_me

# Gerando a planilha de erros
def gerar_planilha_erros(planilha_geral):
    erros = planilha_geral[[
        'codigo_negociacao',
        'valor_empresa',
        'valor_empresas_apoiadas',
        'valor_outras_empresas',
        'valor_unidade_embrapii',
        'valor_unidade_embrapii2',
        'valor_sebrae',
        'valor_sebrae2',
        'valor_embrapii',
        'valor_embrapii2',
        ]]
    
    erros['valor_empresa2'] = erros['valor_empresas_apoiadas'] + erros['valor_outras_empresas']
    
    erros['comp_valor_empresa'] = round(erros['valor_empresa'] - erros['valor_empresa2'], 2)
    erros['comp_valor_unidade_embrapii'] = round(erros['valor_unidade_embrapii'] - erros['valor_unidade_embrapii2'],2)
    erros['comp_valor_sebrae'] = round(erros['valor_sebrae'] - erros['valor_sebrae2'],2)
    erros['comp_valor_embrapii'] = round(erros['valor_embrapii'] - erros['valor_embrapii2'],2)

    erros = erros[
    (erros['comp_valor_empresa'] != 0) |
    (erros['comp_valor_unidade_embrapii'] != 0) |
    (erros['comp_valor_sebrae'] != 0) |
    (erros['comp_valor_embrapii'] != 0)
    ]

    erros = erros.rename(columns={
        'valor_empresa': 'empresa_port',
        'valor_unidade_embrapii': 'unidade_embrapii_port',
        'valor_sebrae': 'sebrae_port',
        'valor_embrapii': 'embrapii_port',
        'valor_empresa2': 'empresa_fonte',
        'valor_unidade_embrapii2': 'unidade_embrapii_fonte',
        'valor_sebrae2': 'sebrae_fonte',
        'valor_embrapii2': 'embrapii_fonte',
        'comp_valor_empresa': 'empresa_port_menos_fonte',
        'comp_valor_unidade_embrapii': 'unidade_embrapii_port_menos_fonte',
        'comp_valor_sebrae': 'sebrae_port_menos_fonte',
        'comp_valor_embrapii': 'embrapii_port_menos_fonte',
        })
    
    # reordenando as colunas
    erros = erros[[
        'codigo_negociacao',
        'empresa_port',
        'empresa_fonte',
        'empresa_port_menos_fonte',
        'unidade_embrapii_port',
        'unidade_embrapii_fonte',
        'unidade_embrapii_port_menos_fonte',
        'sebrae_port',
        'sebrae_fonte',
        'sebrae_port_menos_fonte',
        'embrapii_port',
        'embrapii_fonte',
        'embrapii_port_menos_fonte',
        ]]

    # salvando os erros em STEP3
    erros.to_excel(os.path.abspath(os.path.join(STEP3, 'erros_valores.xlsx')), index = False)


def copiar_planilha_formatada(origem, destino, nome_aba):
    wb_origem = load_workbook(origem)
    wb_destino = load_workbook(destino)

    if nome_aba in wb_destino.sheetnames:
        del wb_destino[nome_aba]

    ws_origem = wb_origem[nome_aba]
    ws_destino = wb_destino.create_sheet(title=nome_aba)

    for row in ws_origem:
        for cell in row:
            new_cell = ws_destino.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)

    # Copiar larguras das colunas
    for col_letter, col_dim in ws_origem.column_dimensions.items():
        ws_destino.column_dimensions[col_letter].width = col_dim.width

    # Mover aba para a primeira posição
    pos = wb_destino.worksheets.index(ws_destino)
    wb_destino.move_sheet(ws_destino, offset=-pos)

    wb_destino.save(destino)
    wb_origem.close()
    wb_destino.close()

