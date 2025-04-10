import os
import shutil
import openpyxl
import sys
from dotenv import load_dotenv

# Carregar .env
load_dotenv()
ROOT = os.getenv('ROOT')
USUARIO = os.getenv('USERNAME')
sys.path.append(ROOT)

# Função para registrar atualizações
def registrar_log(log):

    # Caminhos dos arquivos
    DWPII_COPY = os.path.abspath(os.path.join(ROOT, 'DWPII_copy'))
    DWPII_UP = os.path.abspath(os.path.join(ROOT, 'DWPII_up'))
    log_copy = os.path.join(DWPII_COPY, "historico.xlsx")
    log_up = os.path.join(DWPII_UP, "historico.xlsx")

    # Verificar se o arquivo existe
    if not os.path.exists(log_copy):
        # Criar uma nova planilha e adicionar cabeçalhos
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Log"
        ws.append(["ID", "Data da Atualização", "Usuário", "Nome da Planilha"])
        wb.save(log_copy)

    # Fazer uma cópia para DWPII UP caso exista
    if not os.path.exists(log_up):
        shutil.copy(log_copy, log_up)

    # Abrir o arquivo existente
    wb = openpyxl.load_workbook(log_up)
    ws = wb["Log"]

    # Encontrar o próximo ID
    max_row = ws.max_row
    next_id = max_row if max_row > 1 else 1

    # Adicionar novas entradas de log
    for entry in log:
        ws.append([next_id, entry[0], entry[1], entry[2]])
        next_id += 1

    # Salvar o arquivo
    wb.save(log_up)

# Executar função
if __name__ == "__main__":
    registrar_log()
